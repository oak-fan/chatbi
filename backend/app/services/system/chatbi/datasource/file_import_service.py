"""表格文件导入到 PostgreSQL（CSV / Excel）。"""

from __future__ import annotations

import csv
import re
from collections.abc import Iterator
from contextlib import suppress
from pathlib import Path

from openpyxl import load_workbook

from .....constants.chatbi.datasource import CHATBI_DEFAULT_SCHEMA_NAME
from .....domain.system.chatbi.datasource import (
    MAX_PG_IDENT_LEN,
    build_import_column_names,
    validate_pg_ident,
)
from .connectors.postgresql import PostgreSQLConnector

_SHORT_ID_LEN = 8
_TABULAR_SUFFIXES = frozenset({".csv", ".xlsx"})
_IMPORT_BATCH_SIZE = 500


def _safe_sql_ident(name: str, *, field: str) -> str:
    return validate_pg_ident(name, field=field)


def _quote_pg_ident(name: str) -> str:
    return f'"{name.replace('"', '""')}"'


def slug_sql_ident(raw: str, *, fallback: str = "item", max_len: int = MAX_PG_IDENT_LEN) -> str:
    """将展示名转为 PostgreSQL 安全标识符片段（小写字母、数字、下划线）。"""

    text = raw.strip().lower()
    if not text:
        return fallback
    chars: list[str] = []
    for ch in text:
        if ch.isascii() and ch.isalnum():
            chars.append(ch)
        else:
            chars.append("_")
    cleaned = re.sub(r"_+", "_", "".join(chars)).strip("_")
    if not cleaned:
        cleaned = fallback
    if cleaned[0].isdigit():
        cleaned = f"_{cleaned}"
    if len(cleaned) > max_len:
        cleaned = cleaned[:max_len].rstrip("_")
    if not cleaned or cleaned == "_":
        return fallback
    if cleaned[0].isdigit():
        cleaned = f"_{cleaned[: max(1, max_len - 2)]}".rstrip("_")
    return cleaned or fallback


def format_datasource_short_id(datasource_id: int, *, length: int = _SHORT_ID_LEN) -> str:
    """将雪花 ID 编码为固定长度的 base36 短码。"""

    alphabet = "0123456789abcdefghijklmnopqrstuvwxyz"
    n = max(int(datasource_id), 0)
    if n == 0:
        return "0" * length
    chars: list[str] = []
    while n:
        n, rem = divmod(n, 36)
        chars.append(alphabet[rem])
    code = "".join(reversed(chars))
    if len(code) > length:
        return code[-length:]
    return code.rjust(length, "0")


def _truncate_pg_ident_bytes(name: str, max_bytes: int = MAX_PG_IDENT_LEN) -> str:
    if len(name.encode("utf-8")) <= max_bytes:
        return name
    chars: list[str] = []
    total = 0
    for char in name:
        char_len = len(char.encode("utf-8"))
        if total + char_len > max_bytes:
            break
        chars.append(char)
        total += char_len
    return "".join(chars).rstrip("_") or _GENERIC_IMPORT_TABLE_STEM


def build_file_upload_schema_name(datasource_name: str, datasource_id: int) -> str:
    """表格上传数据源在 PG 中的 schema：{slug(name)}_{id短码}。"""

    short = format_datasource_short_id(datasource_id)
    suffix = f"_{short}"
    max_slug_len = MAX_PG_IDENT_LEN - len(suffix)
    slug = slug_sql_ident(datasource_name, fallback="datasource", max_len=max_slug_len)
    return _safe_sql_ident(f"{slug}{suffix}", field="schema_name")


_GENERIC_IMPORT_TABLE_STEM = "table"


def _import_table_stem(original_name: str) -> str:
    """由上传文件名（无扩展名）得到 PG 双引号表名，保留中文。"""
    path = Path(original_name)
    name = (path.name or "").strip()
    if not name:
        return _GENERIC_IMPORT_TABLE_STEM

    suffix = path.suffix
    stem = (path.stem or "").strip()
    # 无有效主文件名：`.csv`（POSIX 上 suffix 常为空）、整段即扩展名等
    if suffix:
        ext = suffix.lstrip(".")
        if name == suffix or (name.startswith(".") and stem in {ext, suffix}):
            return _GENERIC_IMPORT_TABLE_STEM
    elif name.startswith("."):
        return _GENERIC_IMPORT_TABLE_STEM

    if not stem:
        return _GENERIC_IMPORT_TABLE_STEM
    try:
        return validate_pg_ident(stem, field="表名")
    except ValueError:
        return _GENERIC_IMPORT_TABLE_STEM


def _allocate_table_name(
    stem: str,
    *,
    file_id: int,
    used: set[str],
    reserved: set[str],
) -> str:
    """在同 schema 内分配唯一表名：stem → stem_1…n → stem_{file短码}（通用 stem 从 _1 起）。"""

    base = stem
    candidates: list[str] = []
    if base == _GENERIC_IMPORT_TABLE_STEM:
        for n in range(1, 1000):
            candidates.append(f"{base}_{n}")
    else:
        candidates.append(base)
        for n in range(2, 1000):
            candidates.append(f"{base}_{n}")
    candidates.append(f"{base}_{format_datasource_short_id(file_id)}")

    for raw in candidates:
        raw = _truncate_pg_ident_bytes(raw)
        try:
            name = validate_pg_ident(raw, field="表名")
        except ValueError:
            continue
        if name not in used and name not in reserved:
            used.add(name)
            return name
    msg = "无法分配唯一表名"
    raise ValueError(msg)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _iter_tabular_rows(path: Path) -> Iterator[list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            yield from csv.reader(handle)
        return
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            if worksheet is None:
                msg = "Excel 文件无有效工作表"
                raise ValueError(msg)
            for row in worksheet.iter_rows(values_only=True):
                yield [_cell_text(cell) for cell in row]
        finally:
            workbook.close()
        return
    supported = ", ".join(sorted(_TABULAR_SUFFIXES))
    msg = f"仅支持 {supported} 文件导入"
    raise ValueError(msg)


def _iter_import_values(
    rows: Iterator[list[str]],
    *,
    column_count: int,
) -> Iterator[tuple[str | None, ...]]:
    for row in rows:
        values: list[str | None] = []
        for idx in range(column_count):
            raw = row[idx] if idx < len(row) else ""
            values.append(None if raw == "" else str(raw)[:8000])
        yield tuple(values)


class ChatbiFileImportService:
    """将已下载的表格文件导入到指定 schema。"""

    def __init__(self) -> None:
        self._connector = PostgreSQLConnector()

    async def drop_schema(
        self,
        connector_config: dict[str, object],
        *,
        schema_name: str,
    ) -> None:
        """删除表格上传数据源在 PG 中的隔离 schema（含表）。"""

        schema = _safe_sql_ident(schema_name.strip(), field="schema_name")
        if schema == CHATBI_DEFAULT_SCHEMA_NAME:
            return
        await self._connector.execute_sql_transaction(
            connector_config,
            [f'DROP SCHEMA IF EXISTS "{schema}" CASCADE'],
        )

    async def _existing_table_names(
        self,
        connector_config: dict[str, object],
        schema: str,
    ) -> set[str]:
        schema_esc = schema.replace("'", "''")
        _cols, rows = await self._connector.execute_readonly_sql(
            connector_config,
            (
                "SELECT table_name FROM information_schema.tables "
                f"WHERE table_schema = '{schema_esc}' AND table_type = 'BASE TABLE'"  # nosec
            ),
        )
        return {str(row["table_name"]) for row in rows}

    async def import_files(
        self,
        connector_config: dict[str, object],
        *,
        schema_name: str,
        files: list[tuple[int, Path, str]],
        replace_schema: bool = False,
    ) -> None:
        """在目标库建 schema/表并将多个表格文件灌入（每文件一张表）。"""

        schema = _safe_sql_ident(schema_name, field="schema_name")
        if replace_schema and schema != CHATBI_DEFAULT_SCHEMA_NAME:
            await self._connector.execute_sql_transaction(
                connector_config,
                [f"DROP SCHEMA IF EXISTS {_quote_pg_ident(schema)} CASCADE"],
            )
            reserved: set[str] = set()
        else:
            reserved = await self._existing_table_names(connector_config, schema)
        used: set[str] = set()
        created_tables: list[str] = []

        try:
            for file_id, path, original_name in files:
                if path.suffix.lower() not in _TABULAR_SUFFIXES:
                    supported = ", ".join(sorted(_TABULAR_SUFFIXES))
                    msg = f"仅支持 {supported} 文件导入"
                    raise ValueError(msg)
                table_stem = _import_table_stem(original_name)
                table = _allocate_table_name(
                    table_stem,
                    file_id=file_id,
                    used=used,
                    reserved=reserved,
                )
                rows = _iter_tabular_rows(path)
                try:
                    header = next(rows)
                except StopIteration as exc:
                    msg = "表格文件为空"
                    raise ValueError(msg) from exc
                cols = build_import_column_names(header)
                col_defs = ", ".join(f"{_quote_pg_ident(c)} TEXT" for c in cols)
                table_ref = f"{_quote_pg_ident(schema)}.{_quote_pg_ident(table)}"
                ddl_statements = [
                    f"CREATE SCHEMA IF NOT EXISTS {_quote_pg_ident(schema)}",
                    f"CREATE TABLE {table_ref} ({col_defs})",
                ]
                cols_sql = ", ".join(_quote_pg_ident(c) for c in cols)
                placeholders = ", ".join(f"${idx}" for idx in range(1, len(cols) + 1))
                created_tables.append(table)
                await self._connector.execute_table_import(
                    connector_config,
                    ddl_statements=ddl_statements,
                    insert_sql=(
                        f"INSERT INTO {table_ref} ({cols_sql}) VALUES ({placeholders})"  # nosec
                    ),
                    rows=_iter_import_values(rows, column_count=len(cols)),
                    batch_size=_IMPORT_BATCH_SIZE,
                )
        except Exception:
            cleanup_sql = [
                f"DROP TABLE IF EXISTS {_quote_pg_ident(schema)}.{_quote_pg_ident(table)} CASCADE"
                for table in reversed(created_tables)
                if table not in reserved
            ]
            if cleanup_sql:
                with suppress(Exception):
                    await self._connector.execute_sql_transaction(connector_config, cleanup_sql)
            raise


__all__ = [
    "ChatbiFileImportService",
    "build_file_upload_schema_name",
    "format_datasource_short_id",
    "slug_sql_ident",
]
